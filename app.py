import os
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_from_directory, jsonify
from flask_login import UserMixin, LoginManager, login_user, logout_user, login_required, current_user
from dotenv import load_dotenv
from models import User, Lab, Case, Sample, CustodyEvent, ROLE_ADMIN, ROLE_LAB, ROLE_OFFICER
from security import LoginUser
from utils import compute_priority, make_qr
from werkzeug.security import generate_password_hash, check_password_hash
import csv
from openpyxl import load_workbook
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import secrets
from storage import Storage
import pandas as pd
from flask_login import login_user


EXCEL_FILE = 'forensic_cases.xlsx'
xls = pd.ExcelFile(EXCEL_FILE)
print(xls.sheet_names)

load_dotenv()
SECRET_KEY = os.getenv('SECRET_KEY', 'dev')
ADMIN_EMAIL = os.getenv('ADMIN_EMAIL', 'admin@fasttrack.local')
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'ChangeMe123!')
USE_EXCEL = bool(int(os.getenv('USE_EXCEL','1')))

app = Flask(__name__, static_folder='static', template_folder='templates')
app.config['SECRET_KEY'] = SECRET_KEY

login_manager = LoginManager(app)
login_manager.login_view = 'login'

storage = Storage('forensic_cases.xlsx')

# simple user wrapper
class WebUser(UserMixin):
    def __init__(self, row):
        self.id = str(row.get('id'))
        self.email = row.get('email')
        self.name = row.get('name')
        self.role = row.get('role')
        self.api_token = row.get('api_token')

    @property
    def is_admin(self):
        return self.role == 'admin'

    @property
    def is_lab(self):
        return self.role == 'lab'

    @property
    def is_officer(self):
        return self.role == 'officer'

@login_manager.user_loader
def load_user(user_id):
    u = storage.find('users', id=str(user_id))
    if not u:
        return None
    return WebUser(u)

def bootstrap_admin():
    users = storage.all('users')
    if not users:
        hashed = generate_password_hash(ADMIN_PASSWORD)
        admin = {
            'email': ADMIN_EMAIL,
            'name': 'Admin',
            'role': 'admin',
            'password_hash': hashed,
            'api_token': '',
        }
        storage.append('users', admin)
bootstrap_admin()

def hash_password(password):
    return generate_password_hash(password)

# check password during login
def check_password_hash_stored(stored_hash, password):
    return check_password_hash(stored_hash, password)

# ---------- UTILITIES ----------
def read_sheet(sheet_name):
    df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
    return df

def write_excel(sheet, df):
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet, index=False)

def append_sheet(sheet_name, data_dict):
    df = read_sheet(sheet_name)
    df = pd.concat([df, pd.DataFrame([data_dict])], ignore_index=True)
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def generate_qr(data):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill="black", back_color="white")
    buf = BytesIO()
    img.save(buf)
    buf.seek(0)
    return buf

def export_case_pdf(case_id):
    df_cases = read_sheet("Cases")
    case = df_cases[df_cases["CaseID"] == case_id].iloc[0]
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, f"Case Report - {case_id}", ln=True)
    pdf.set_font("Arial", "", 12)
    for col in df_cases.columns:
        pdf.cell(0, 10, f"{col}: {case[col]}", ln=True)
    pdf_file = BytesIO()
    pdf.output(pdf_file)
    pdf_file.seek(0)
    return pdf_file

def calculate_priority(status, suspect):
    score = 0
    if "urgent" in status.lower() or "high-risk" in suspect.lower():
        score += 10
    if "crime scene" in status.lower():
        score += 5
    return score

def get_priority_tuple(status, suspect):
    result = calculate_priority(status, suspect)
    if isinstance(result, (list, tuple)) and len(result) == 2:
        return result
    else:
        return (result, None)

# Helper Functions
# -------------------

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        with pd.ExcelWriter(EXCEL_FILE) as writer:
            pd.DataFrame(columns=['CaseID','Suspect','Status','Priority','CreatedAt']).to_excel(writer,sheet_name='Cases', index=False)
            pd.DataFrame(columns=['CaseID','SampleID','Type','CollectedBy','CollectedAt']).to_excel(writer,sheet_name='Samples', index=False)
            pd.DataFrame(columns=['CaseID','Event','Timestamp']).to_excel(writer,sheet_name='ChainOfCustody', index=False)

def read_sheet(sheet_name):
    return pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)

def write_sheet(df, sheet_name):
    book = load_workbook(EXCEL_FILE)
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()

def generate_qr(data):
    img = qrcode.make(data)
    buffer = BytesIO()
    img.save(buffer, 'PNG')
    buffer.seek(0)
    return buffer

def export_case_pdf(case_id):
    cases = read_sheet('Cases')
    samples = read_sheet('Samples')
    coc = read_sheet('ChainOfCustody')
    
    case = cases[cases['CaseID']==case_id].iloc[0]
    case_samples = samples[samples['CaseID']==case_id]
    case_coc = coc[coc['CaseID']==case_id]

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(0,10,f"Case Report: {case_id}", ln=True)
    pdf.set_font("Arial", '', 12)
    pdf.cell(0,10,f"Suspect: {case['Suspect']}", ln=True)
    pdf.cell(0,10,f"Status: {case['Status']}", ln=True)
    pdf.cell(0,10,f"Priority: {case['Priority']}", ln=True)
    pdf.cell(0,10,f"Created At: {case['CreatedAt']}", ln=True)

    pdf.ln(5)
    pdf.cell(0,10,"Samples:", ln=True)
    for _, s in case_samples.iterrows():
        pdf.cell(0,8,f"{s['SampleID']} | {s['Type']} | Collected by {s['CollectedBy']} on {s['CollectedAt']}", ln=True)

    pdf.ln(5)
    pdf.cell(0,10,"Chain of Custody:", ln=True)
    for _, e in case_coc.iterrows():
        pdf.cell(0,8,f"{e['Timestamp']} - {e['Event']}", ln=True)

    # Save PDF
    pdf_file = f"{case_id}_report.pdf"
    pdf.output(pdf_file)
    return pdf_file


# -------- Auth --------
@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        pwd = request.form['password']
        user = storage.find('users', email=email)
        if not user:
            flash('Invalid credentials', 'danger')
            return redirect(url_for('login'))
        if check_password_hash_stored(user.get('password_hash',''), pwd):
            u = WebUser(user)
            login_user(u)
            return redirect(url_for('dashboard'))
        flash('Invalid credentials', 'danger')
        return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        password = request.form['password']

        # Check if user exists
        users = storage.all('users')
        if any(u.get('email') == email for u in users):
            flash("Email already exists", "danger")
            return redirect(url_for('register'))

        # Save new user
        new_user = {
            "email": email,
            "name": email.split('@')[0],  # optional default name
            "role": "officer",  # default role
            "password_hash": generate_password_hash(password),
            "api_token": ""
        }
        storage.append('users', new_user)

        flash("Registration successful. You can now log in.", "success")
        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/')
@login_required
def dashboard():
    df = pd.read_excel(EXCEL_FILE, sheet_name="cases")  # ⚠ capital C, check Excel

    def priority_wrapper(status, suspect):
        result = calculate_priority(status, suspect)
        if isinstance(result, (list, tuple)):
            if len(result) >= 2:
                return (result[0], result[1])
            elif len(result) == 1:
                return (result[0], None)
            else:
                return (None, None)
        else:
            return (result, None)

    # Build a list of tuples [(priority, label), ...]
    priority_data = [priority_wrapper(row["Status"], row["Suspect"]) for _, row in df.iterrows()]
    
    # Create a DataFrame with exactly 2 columns
    priority_df = pd.DataFrame(priority_data, columns=["Priority", "PriorityLabel"], index=df.index)

    # Merge into df
    df = pd.concat([df, priority_df], axis=1)

    print(df[["Status", "Suspect", "Priority", "PriorityLabel"]].head())  # debug

    df_sorted = df.sort_values(by="Priority", ascending=False)
    cases = storage.all('cases')

    def sort_key(c):
        try:
            return (-int(c.get('priority_score', 0)), c.get('created_at', ''))
        except:
            return (0, c.get('created_at', ''))

    cases_sorted = sorted(cases, key=sort_key)[:200]
    return render_template('dashboard.html', cases=cases_sorted)


@app.route('/admin/users', methods=['GET','POST'])
@login_required
def users_list():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form['email'].strip().lower()
        name = request.form['name']
        role = request.form['role']
        pwd = request.form['password']
        api_token = request.form.get('api_token','').strip()
        if not api_token:
            api_token = ''
        hashed = hash_password(pwd)
        storage.append('users', {'email': email, 'name': name, 'role': role, 'password_hash': hashed, 'api_token': api_token})
        flash('User added', 'success')
        return redirect(url_for('users_list'))
    users = storage.all('users')
    return render_template('users_list.html', users=users)

@app.route("/send_to_lab/<case_id>")
@login_required
def send_to_lab(case_id):
    df = pd.read_excel(EXCEL_FILE, sheet_name="Cases")
    idx = df.index[df["CaseID"] == case_id].tolist()
    if not idx:
        flash("Case not found")
        return redirect(url_for("dashboard"))
    idx = idx[0]

    # Simulate lab processing
    import random, time
    time.sleep(1)  # simulate processing delay
    results = random.choice(["DNA matched", "No match found", "Pending further analysis"])
    df.at[idx, "Status"] = f"Lab Result: {results}"
    df.to_excel(EXCEL_FILE, sheet_name="Cases", index=False)
    flash(f"Lab results updated for Case {case_id}")
    return redirect(url_for("dashboard"))

@app.route('/admin/labs', methods=['GET','POST'])
@login_required
def labs_list():
    if not current_user.is_admin:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        name = request.form['name']
        contact = request.form.get('contact_email','')
        storage.append('labs', {'name': name, 'contact_email': contact, 'is_active': True})
        flash('Lab added', 'success')
        return redirect(url_for('labs_list'))
    labs = storage.all('labs')
    return render_template('labs_list.html', labs=labs)

@app.route('/cases/new', methods=['GET','POST'])
@login_required
def case_new():
    if request.method == 'POST':
        case_number = request.form['case_number'].strip()
        offence_type = request.form['offence_type']
        description = request.form.get('description','')
        in_custody = request.form.get('suspect_in_custody') == 'on'
        priority = compute_priority(offence_type, age_days=0, suspect_in_custody=in_custody)
        c = {'case_number': case_number, 'offence_type': offence_type, 'description': description, 'priority_score': priority, 'status': 'created', 'created_by': current_user.email}
        new_case = storage.append('cases', c)
        # create default sample
        code = f"S-{new_case['id']:06d}-A"
        qr_path = make_qr(code)
        s = {'case_number': case_number, 'code': code, 'qr_path': qr_path, 'status': 'sealed'}
        storage.append('samples', s)
        # custody event
        prev_hash = storage.last_event_hash(case_number)
        payload = {'actor': current_user.email, 'action': 'created_case', 'sample_code': code, 'timestamp': datetime.utcnow().isoformat()}
        h = compute_event_hash(prev_hash, payload)
        ev = {'case_number': case_number, 'sample_code': code, 'actor': current_user.email, 'action': 'created_case', 'timestamp': datetime.utcnow().isoformat(), 'note': 'Case and sample created', 'prev_hash': prev_hash, 'hash': h}
        storage.append('custody_events', ev)
        flash('Case created', 'success')
        return redirect(url_for('case_detail', case_number=case_number))
    return render_template('case_new.html')

@app.route('/cases/<case_number>')
@login_required
def case_detail(case_number):
    case = storage.find('cases', case_number=case_number)
    if not case:
        flash('Case not found', 'danger')
        return redirect(url_for('dashboard'))
    samples = storage.filter('samples', case_number=case_number)
    events = storage.filter('custody_events', case_number=case_number)
    events_sorted = sorted(events, key=lambda e: e.get('timestamp',''))
    return render_template('case_detail.html', case=case, samples=samples, events=events_sorted)

@app.route('/cases/<case_number>/status', methods=['POST'])
@login_required
def case_status(case_number):
    new_status = request.form['status']
    storage.update('cases', 'case_number', case_number, {'status': new_status})
    # log event
    prev_hash = storage.last_event_hash(case_number)
    payload = {'actor': current_user.email, 'action': f'status:{new_status}', 'timestamp': datetime.utcnow().isoformat()}
    h = compute_event_hash(prev_hash, payload)
    ev = {'case_number': case_number, 'sample_code': '', 'actor': current_user.email, 'action': f'status:{new_status}', 'timestamp': datetime.utcnow().isoformat(), 'note': '', 'prev_hash': prev_hash, 'hash': h}
    storage.append('custody_events', ev)
    flash('Status updated', 'success')
    return redirect(url_for('case_detail', case_number=case_number))

# API for partner labs
def authorize_api(token):
    if not token:
        return None
    u = storage.find('users', api_token=token)
    if not u or u.get('role')!='lab':
        return None
    return u

@app.route('/case/<case_id>', endpoint='case_detail_api')
def case_detail_api(case_id):
    cases = read_sheet('Cases')
    samples = read_sheet('Samples')
    coc = read_sheet('ChainOfCustody')
    case = cases[cases['CaseID']==case_id].iloc[0]
    case_samples = samples[samples['CaseID']==case_id].to_dict(orient='records')
    case_coc = coc[coc['CaseID']==case_id].to_dict(orient='records')
    qr = generate_qr(case_id)
    qr_data = qr.getvalue().hex()
    return render_template('case_detail.html', case=case, samples=case_samples, coc=case_coc)

@app.route("/export_pdf/<case_id>")
def export_pdf(case_id):
    samples = load_sheet("Samples")
    chain = load_sheet("ChainOfCustody")
    lab = load_sheet("LabResults")
    case_samples = samples[samples["case_id"]==case_id].to_dict(orient="records")
    case_chain = chain[chain["case_id"]==case_id].to_dict(orient="records")
    case_lab = lab[lab["case_id"]==case_id].to_dict(orient="records")
    rendered = render_template("report.html", case_id=case_id, samples=case_samples, chain=case_chain, lab=case_lab)
    pdf_path = f"reports/{case_id}.pdf"
    os.makedirs("reports", exist_ok=True)
    pdfkit.from_string(rendered, pdf_path)
    return send_file(pdf_path)

@app.route("/add_sample/<case_id>", methods=["GET","POST"])
def add_sample(case_id):
    if "user" not in session:
        return redirect(url_for("index"))
    if request.method=="POST":
        sample_id = request.form["sample_id"]
        desc = request.form["description"]
        qr_path = f"static/qrcodes/{sample_id}.png"
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(f"SampleID:{sample_id}")
        qr.make(fit=True)
        img = qr.make_image(fill="black", back_color="white")
        os.makedirs("static/qrcodes", exist_ok=True)
        img.save(qr_path)
        samples = load_sheet("Samples")
        samples = pd.concat({"sample_id": sample_id, "case_id": case_id, "description": desc, "qr_path": qr_path}, ignore_index=True)
        save_sheet(samples, "Samples")
        return redirect(url_for("case_detail", case_id=case_id))
    return render_template("add_sample.html", case_id=case_id)

@app.route("/add_event/<case_id>", methods=["POST"])
def add_event(case_id):
    if "user" not in session:
        return redirect(url_for("index"))
    event = request.form["event"]
    import datetime
    chain = load_sheet("ChainOfCustody")
    chain = pd.concat({"event_id": len(chain)+1, "case_id": case_id, "event": event, "timestamp": datetime.datetime.now()}, ignore_index=True)
    save_sheet(chain, "ChainOfCustody")
    return redirect(url_for("case_detail", case_id=case_id))

@app.route("/add_chain/<case_id>", methods=["POST"])
def add_chain(case_id):
    event = request.form["event"]
    officer = request.form["officer"]
    append_sheet("ChainOfCustody", {"CaseID": case_id, "Event": event, "Officer": officer})
    flash(f"Chain of custody updated for case {case_id}")
    return redirect(url_for("case_detail", case_id=case_id))

@app.route("/add_lab/<case_id>", methods=["POST"])
def add_lab(case_id):
    if "user" not in session:
        return redirect(url_for("index"))
    lab_name = request.form["lab_name"]
    result = request.form["result"]
    import datetime
    lab = load_sheet("LabResults")
    lab = pd.concat({"case_id": case_id, "lab_name": lab_name, "result": result, "date": datetime.datetime.now()}, ignore_index=True)
    save_sheet(lab, "LabResults")
    return redirect(url_for("case_detail", case_id=case_id))

@app.route('/api/v1/cases/<case_number>/receive', methods=['POST'])
def api_receive(case_number):
    token = request.headers.get('X-API-Token')
    user = authorize_api(token)
    if not user:
        return jsonify({'error':'unauthorized'}), 401
    c = storage.find('cases', case_number=case_number)
    if not c:
        return jsonify({'error':'not found'}), 404
    storage.update('cases', 'case_number', case_number, {'status': 'in_lab', 'lab_assigned': user.get('email')})
    prev_hash = storage.last_event_hash(case_number)
    payload = {'actor': user.get('email'), 'action': 'received_by_lab', 'timestamp': datetime.utcnow().isoformat()}
    h = compute_event_hash(prev_hash, payload)
    ev = {'case_number': case_number, 'sample_code': '', 'actor': user.get('email'), 'action': 'received_by_lab', 'timestamp': datetime.utcnow().isoformat(), 'note': 'Received via API', 'prev_hash': prev_hash, 'hash': h}
    storage.append('custody_events', ev)
    return jsonify({'ok': True})

@app.route('/api/v1/cases/<case_number>/complete', methods=['POST'])
def api_complete(case_number):
    token = request.headers.get('X-API-Token')
    user = authorize_api(token)
    if not user:
        return jsonify({'error':'unauthorized'}), 401
    c = storage.find('cases', case_number=case_number)
    if not c:
        return jsonify({'error':'not found'}), 404
    # store a simple result summary if provided
    result_summary = request.json.get('result_summary') if request.is_json else request.form.get('result_summary','')
    storage.update('cases', 'case_number', case_number, {'status': 'completed'})
    # append lab_result
    storage.append('lab_results', {'case_number': case_number, 'sample_code': '', 'lab_user': user.get('email'), 'result_summary': result_summary, 'result_file': ''})
    prev_hash = storage.last_event_hash(case_number)
    payload = {'actor': user.get('email'), 'action': 'completed_by_lab', 'timestamp': datetime.utcnow().isoformat()}
    h = compute_event_hash(prev_hash, payload)
    ev = {'case_number': case_number, 'sample_code': '', 'actor': user.get('email'), 'action': 'completed_by_lab', 'timestamp': datetime.utcnow().isoformat(), 'note': 'Completed via API', 'prev_hash': prev_hash, 'hash': h}
    storage.append('custody_events', ev)
    return jsonify({'ok': True})

# PDF report generator
@app.route('/cases/<case_number>/report')
@login_required
def case_report(case_number):
    case = storage.find('cases', case_number=case_number)
    if not case:
        flash('Case not found', 'danger')
        return redirect(url_for('dashboard'))
    events = storage.filter('custody_events', case_number=case_number)
    results = storage.filter('lab_results', case_number=case_number)
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, height - 40, f"DNA FastTrack — Case Report: {case_number}")
    c.setFont("Helvetica", 10)
    c.drawString(40, height - 60, f"Offence: {case.get('offence_type','')}")
    c.drawString(40, height - 75, f"Status: {case.get('status','')}")
    c.drawString(40, height - 90, f"Created: {case.get('created_at','')}")
    y = height - 120
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "Chain of Custody:")
    y -= 16
    c.setFont("Helvetica", 9)
    for ev in sorted(events, key=lambda e: e.get('timestamp','')):
        line = f"{ev.get('timestamp','')} — {ev.get('actor','')} — {ev.get('action','')} — {ev.get('note','')}"
        c.drawString(40, y, line[:120])
        y -= 12
        if y < 80:
            c.showPage()
            y = height - 40
    if results:
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y-10, "Lab Results:")
        y -= 26
        c.setFont("Helvetica", 9)
        for r in results:
            line = f"{r.get('created_at','')} — {r.get('lab_user','')} — {r.get('result_summary','')}"
            c.drawString(40, y, line[:120])
            y -= 12
            if y < 80:
                c.showPage()
                y = height - 40
    c.showPage()
    c.save()
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', download_name=f"{case_number}_report.pdf", as_attachment=True)

@app.route("/generate_report/<case_id>")
@login_required
def generate_report(case_id):
    df = pd.read_excel(EXCEL_FILE, sheet_name="Cases")
    case = df[df["CaseID"] == case_id].to_dict(orient="records")
    if not case:
        flash("Case not found")
        return redirect(url_for("dashboard"))
    case = case[0]

    pdf_file = BytesIO()
    c = canvas.Canvas(pdf_file)
    c.drawString(100, 750, f"Case Report: {case['CaseID']}")
    c.drawString(100, 730, f"Suspect: {case['Suspect']}")
    c.drawString(100, 710, f"Status: {case['Status']}")
    c.save()
    pdf_file.seek(0)
    return send_file(pdf_file, as_attachment=True, download_name=f"Case_{case_id}_Report.pdf")

if __name__ == "__main__":
    # Create Excel if not exist
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        wb.create_sheet("Cases", 0)
        ws = wb["Cases"]
        ws.append(["CaseID", "Suspect", "Status"])
        wb.save(EXCEL_FILE)
    app.run(debug=True)